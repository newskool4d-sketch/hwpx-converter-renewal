"""
HWP COM 자동화로 Markdown / TXT / DOCX / HTML / CSV / XLSX / PDF → HWPX 변환.
확장자를 자동 감지하여 내부 blocks 구조로 정규화한 뒤 HWP COM으로 저장.
"""
from pathlib import Path
import argparse
import csv
import json
import os
import re
import subprocess
import sys
import time
import zipfile
import tempfile
import shutil
import xml.etree.ElementTree as ET
from collections.abc import Callable

from hwpx_layout import (
    OFFICIAL_LIST_MAX_DEPTH,
    TABLE_AFTER_PARA_SPACE,
    TABLE_CELL_MARGIN,
    TABLE_CELL_PARA_SPACE,
    TABLE_MIN_ROW_HEIGHT,
    TABLE_TOTAL_WIDTH,
    _profile_to_widths,
    _table_header_profile,
    calc_col_widths,
    calc_row_heights,
    calc_table_after_para_space,
    calc_table_cell_margin,
    calc_table_cell_para_space,
    detect_official_list_item,
    normalize_table_rows,
    official_list_para_shape,
)
from table_grid import SourceCell, block_rows_from_grid, expand_spanned_rows
from table_hwpx_postprocess import apply_table_layout_profiles as _apply_table_layout_profiles_new
from table_model import table_layout_for


DiagnoseStageReporter = Callable[[str], None]


def _configure_utf8_stdio():
    os.environ.setdefault('PYTHONIOENCODING', 'utf-8')
    for stream_name in ('stdout', 'stderr'):
        stream = getattr(sys, stream_name, None)
        if stream is not None and hasattr(stream, 'reconfigure'):
            stream.reconfigure(encoding='utf-8', errors='replace')


def _utf8_subprocess_env():
    env = os.environ.copy()
    env.setdefault('PYTHONIOENCODING', 'utf-8')
    env.setdefault('PYTHONUTF8', '1')
    return env


_configure_utf8_stdio()


SUPPORTED_EXTENSIONS = {'.md', '.txt', '.docx', '.html', '.htm', '.csv', '.xlsx', '.pdf'}
OPTIONAL_DEPENDENCIES = {
    '.docx': 'python-docx',
    '.html': 'beautifulsoup4',
    '.htm': 'beautifulsoup4',
    '.xlsx': 'openpyxl',
    '.pdf': 'opendataloader-pdf (권장), KORDOC_HOME 환경변수 또는 --kordoc-home, pdfplumber, PyMuPDF 또는 pypdf',
}
OUTPUT_MANIFEST_NAME = '.anyway_to_hwpx_output.json'


def as_path(value):
    return Path(value).expanduser().resolve()


def read_text_file(path):
    for encoding in ('utf-8-sig', 'utf-8', 'cp949'):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
    raise UnicodeDecodeError('unknown', b'', 0, 1, f'텍스트 인코딩을 판별할 수 없음: {path}')


def require_file(path):
    if not path.exists():
        raise FileNotFoundError(f'입력 파일을 찾을 수 없음: {path}')
    if not path.is_file():
        raise ValueError(f'입력 경로가 파일이 아님: {path}')


def resolve_kordoc_dir(explicit_path=None):
    candidates = []
    if explicit_path:
        candidates.append(Path(explicit_path))
    for env_name in ('KORDOC_HOME', 'KORDOC_AI_HOME'):
        value = os.environ.get(env_name)
        if value:
            candidates.append(Path(value))
    candidates.append(Path(r'C:/Users/홍주형/kordoc-ai'))

    for candidate in candidates:
        expanded = candidate.expanduser()
        if expanded.exists():
            return expanded.resolve()
    return None


def _kordoc_commands(kordoc_dir, path):
    return [
        ['python', str(kordoc_dir / 'main.py'), str(path)],
        ['python', '-m', 'kordoc', str(path)],
    ]


# ─── Markdown 파서 ─────────────────────────────────────────────────────────────

def _clean_inline(text):
    text = re.sub(r'\[([^\]]+)\]\([^\)]+\)', r'\1', text)
    text = re.sub(r'!\[[^\]]*\]\([^\)]+\)', '', text)
    text = re.sub(r'`([^`]+)`', r'\1', text)
    text = re.sub(r'\*\*([^*]+)\*\*', r'\1', text)
    text = re.sub(r'__([^_]+)__', r'\1', text)
    text = re.sub(r'\*([^*]+)\*', r'\1', text)
    text = re.sub(r'_([^_]+)_', r'\1', text)
    text = text.replace('&nbsp;', ' ')
    text = re.sub(r'<[^>]+>', '', text)
    return text.strip()


def _is_separator(line):
    if len(line) > 500:
        return False
    cells = _split_markdown_table_cells(line)
    return len(cells) >= 1 and all(re.match(r'^[ \t]*:?-+:?[ \t]*$', c) for c in cells)


def _split_markdown_table_cells(line):
    stripped = line.strip()
    if stripped.startswith('|'):
        stripped = stripped[1:]
    if stripped.endswith('|'):
        stripped = stripped[:-1]

    cells = []
    current = []
    escaped = False
    for char in stripped:
        if escaped:
            if char == '|':
                current.append('|')
            else:
                current.append('\\')
                current.append(char)
            escaped = False
            continue
        if char == '\\':
            escaped = True
            continue
        if char == '|':
            cells.append(''.join(current))
            current = []
            continue
        current.append(char)
    if escaped:
        current.append('\\')
    cells.append(''.join(current))
    return cells


def _parse_table_row(line):
    return [_clean_inline(c.strip()) for c in _split_markdown_table_cells(line)]


def _normalize_parsed_table(header, rows):
    normalized, _ = normalize_table_rows(header, rows)
    if not normalized:
        return header, rows
    return normalized[0], normalized[1:]


# 항목체계 로마숫자 최상위 레벨 허용 여부(정본 §2-1).
# True  = 계획서·보고서 관행(Ⅰ. 최상위, 기본값)
# False = 대외 시행문(1.이 최상위, 로마숫자 미인식)
_ALLOW_ROMAN_LEVEL = True


def _detect_list_item(line):
    return detect_official_list_item(line, _clean_inline, allow_roman=_ALLOW_ROMAN_LEVEL)


_ATTACHMENT_HEAD_PATTERN = re.compile(r'^붙임\s*[::]?\s+(\S.*)$')
_ATTACHMENT_ITEM_PATTERN = re.compile(r'^\d{1,2}\.\s+.*\d+\s*(?:부|매|장|권|개|식|건)\.?\s*$')


def _detect_attachment_head(line):
    """'붙임  1. ○○계획서 1부.' 형식 감지 → 붙임 뒤 2타로 정규화."""
    match = _ATTACHMENT_HEAD_PATTERN.match(line.strip())
    if match is None:
        return None
    return {'type': 'attachment', 'text': '붙임  ' + match.group(1).strip()}


def _bullet_depth(raw_line, last_official_depth, heading_level=0):
    """공문서 위계: 불릿(•)은 직전 항목기호의 한 단계 하위 들여쓰기를 따른다.

    마크다운 들여쓰기(2칸 = 1단계)는 추가 하위 단계로 누적된다.
    직전 항목기호가 없으면 제목 수준 → 기본값(1) 순서로 기준을 잡는다.
    """
    expanded = raw_line.replace('\t', '  ')
    indent_levels = (len(expanded) - len(expanded.lstrip(' '))) // 2
    if last_official_depth is not None:
        base = last_official_depth + 1
    elif heading_level > 0:
        base = heading_level
    else:
        base = 1
    return min(base + indent_levels, OFFICIAL_LIST_MAX_DEPTH)


def parse_markdown(text):
    lines = text.splitlines()
    blocks = []
    i = 0
    in_front = False
    current_heading_level = 0  # tracks last seen # heading depth
    last_official_depth = None  # depth of the most recent official-numbered list item
    in_attachment = False  # 붙임 목록 연속 항목 추적

    while i < len(lines):
        line = lines[i]

        if not line.strip():
            in_attachment = False
            i += 1
            continue

        if line.strip() == '---':
            if i == 0:
                in_front = True
                i += 1
                continue
            elif in_front:
                in_front = False
                i += 1
                continue
            else:
                blocks.append({'type': 'hr'})
                i += 1
                continue

        if in_front:
            i += 1
            continue

        # 붙임 표시 (시행규칙 제4조제4항): 첫 줄 '붙임  1. …', 이후 '2. ○○ 1부.' 연속 항목
        attachment = _detect_attachment_head(line)
        if attachment:
            in_attachment = True
            blocks.append(attachment)
            i += 1
            continue
        if in_attachment:
            if _ATTACHMENT_ITEM_PATTERN.match(line.strip()):
                blocks.append({'type': 'attachment', 'text': line.strip(), 'cont': True})
                i += 1
                continue
            in_attachment = False

        stripped_line = line.strip()

        if re.match(r'^(수신|경유|제목)\s*:', stripped_line):
            colon_idx = stripped_line.index(':')
            key = stripped_line[:colon_idx].strip()
            value = _clean_inline(stripped_line[colon_idx + 1:].strip())
            blocks.append({'type': 'official_header', 'key': key, 'value': value})
            i += 1
            continue

        if re.match(r'^-{3,}\s*$', line) or re.match(r'^\*{3,}\s*$', line):
            blocks.append({'type': 'hr'})
            i += 1
            continue

        m = re.match(r'^(#{1,3})\s+(.*)', line)
        if m:
            heading_level = len(m.group(1))
            heading_text = m.group(2)
            current_heading_level = heading_level
            # Convert headings whose text matches official Korean numbering (1. 가. etc.) to list items.
            # Plain headings (no numbered prefix) remain as h blocks.
            li_result = _detect_list_item(heading_text)
            if li_result and li_result['marker'] != '•':
                last_official_depth = li_result['depth']
                blocks.append({
                    'type': 'li',
                    'text': li_result['text'],
                    'depth': li_result['depth'],
                    'marker': li_result['marker'],
                    'content': li_result['content'],
                })
            else:
                last_official_depth = None
                blocks.append({'type': 'h', 'level': heading_level, 'text': _clean_inline(heading_text)})
            i += 1
            continue

        if line.strip().startswith('|') and i + 1 < len(lines) and _is_separator(lines[i + 1]):
            header = _parse_table_row(line)
            i += 2
            rows = []
            while i < len(lines) and lines[i].strip().startswith('|'):
                rows.append(_parse_table_row(lines[i]))
                i += 1
            header, rows = _normalize_parsed_table(header, rows)
            blocks.append({'type': 'table', 'header': header, 'rows': rows, 'table_source': 'markdown'})
            continue

        li_result = _detect_list_item(line)
        if li_result:
            depth = li_result['depth']
            if li_result['marker'] == '•':
                depth = _bullet_depth(line, last_official_depth, current_heading_level)
            else:
                last_official_depth = depth
            blocks.append({
                'type': 'li',
                'text': li_result['text'],
                'depth': depth,
                'marker': li_result['marker'],
                'content': li_result['content'],
            })
            i += 1
            continue

        if line.strip().startswith('>'):
            text = re.sub(r'^>\s*', '', line.strip())
            if text:
                blocks.append({'type': 'bq', 'text': _clean_inline(text)})
            i += 1
            continue

        if line.strip().startswith('```'):
            i += 1
            code_lines = []
            while i < len(lines) and not lines[i].strip().startswith('```'):
                code_lines.append(lines[i])
                i += 1
            i += 1
            for cl in code_lines:
                if cl.strip():
                    blocks.append({'type': 'code', 'text': cl})
            continue

        t = _clean_inline(line)
        if t:
            blocks.append({'type': 'p', 'text': t})
        i += 1

    return blocks


def _split_tab_row(line):
    return [cell.strip() for cell in line.rstrip('\n').split('\t')]


def _tab_table_run(lines, start):
    """start부터 탭 구분 표 행이 몇 줄 이어지는지 센다 (2열 이상, 2줄 이상)."""
    count = 0
    while start + count < len(lines):
        line = lines[start + count]
        if '\t' not in line or len(_split_tab_row(line)) < 2 or not line.strip():
            break
        count += 1
    return count if count >= 2 else 0


def _pipe_table_run(lines, start):
    """start부터 파이프(|) 표 행이 몇 줄 이어지는지 센다 (2줄 이상)."""
    count = 0
    while start + count < len(lines) and lines[start + count].strip().startswith('|'):
        count += 1
    return count if count >= 2 else 0


def parse_plain_text(text):
    lines = text.splitlines()
    blocks = []
    last_official_depth = None
    in_attachment = False
    i = 0
    while i < len(lines):
        raw_line = lines[i]
        line = raw_line.strip()
        if not line:
            in_attachment = False
            i += 1
            continue

        # 표 구조 보존: 탭 구분 표 (TSV 형태)
        tab_run = _tab_table_run(lines, i)
        if tab_run:
            rows = [_split_tab_row(lines[i + k]) for k in range(tab_run)]
            header, body = _normalize_parsed_table(rows[0], rows[1:])
            blocks.append({'type': 'table', 'header': header, 'rows': body, 'table_source': 'text'})
            i += tab_run
            continue

        # 표 구조 보존: 파이프(|) 구분 표
        pipe_run = _pipe_table_run(lines, i)
        if pipe_run:
            table_lines = [lines[i + k] for k in range(pipe_run)]
            if len(table_lines) >= 2 and _is_separator(table_lines[1]):
                header = _parse_table_row(table_lines[0])
                body_rows = [_parse_table_row(ln) for ln in table_lines[2:]]
            else:
                header = _parse_table_row(table_lines[0])
                body_rows = [_parse_table_row(ln) for ln in table_lines[1:]]
            header, body_rows = _normalize_parsed_table(header, body_rows)
            blocks.append({'type': 'table', 'header': header, 'rows': body_rows, 'table_source': 'text'})
            i += pipe_run
            continue

        # 붙임 표시 (시행규칙 제4조제4항)
        attachment = _detect_attachment_head(line)
        if attachment:
            in_attachment = True
            blocks.append(attachment)
            i += 1
            continue
        if in_attachment:
            if _ATTACHMENT_ITEM_PATTERN.match(line):
                blocks.append({'type': 'attachment', 'text': line, 'cont': True})
                i += 1
                continue
            in_attachment = False

        li_result = _detect_list_item(line)
        if li_result:
            depth = li_result['depth']
            if li_result['marker'] == '•':
                depth = _bullet_depth(raw_line, last_official_depth)
            else:
                last_official_depth = depth
            blocks.append({
                'type': 'li',
                'text': li_result['text'],
                'depth': depth,
                'marker': li_result['marker'],
                'content': li_result['content'],
            })
        else:
            blocks.append({'type': 'p', 'text': _clean_inline(line)})
        i += 1
    return blocks


def parse_html(text):
    try:
        from bs4 import BeautifulSoup
    except ImportError as exc:
        raise RuntimeError('HTML 변환에는 beautifulsoup4가 필요함: pip install beautifulsoup4') from exc

    soup = BeautifulSoup(text, 'html.parser')
    blocks = []
    body = soup.body or soup
    for node in body.find_all(['h1', 'h2', 'h3', 'p', 'blockquote', 'pre', 'table', 'li'], recursive=True):
        if node.find_parent(['table']) and node.name != 'table':
            continue
        if node.find_parent(['blockquote', 'pre']) and node.name not in ('blockquote', 'pre'):
            continue
        if node.find_parent('li') and node.name != 'li':
            continue
        if node.name in ('h1', 'h2', 'h3'):
            level = int(node.name[1])
            value = node.get_text(' ', strip=True)
            if value:
                blocks.append({'type': 'h', 'level': level, 'text': _clean_inline(value)})
        elif node.name == 'p':
            value = node.get_text(' ', strip=True)
            if value:
                blocks.append({'type': 'p', 'text': _clean_inline(value)})
        elif node.name == 'blockquote':
            value = node.get_text(' ', strip=True)
            if value:
                blocks.append({'type': 'bq', 'text': _clean_inline(value)})
        elif node.name == 'pre':
            value = node.get_text('\n', strip=True)
            for line in value.splitlines():
                if line.strip():
                    blocks.append({'type': 'code', 'text': line.rstrip()})
        elif node.name == 'li':
            parts = []
            for child in node.contents:
                if getattr(child, 'name', None) in ('ul', 'ol'):
                    continue
                if hasattr(child, 'get_text'):
                    child_text = child.get_text(' ', strip=True)
                else:
                    child_text = str(child).strip()
                if child_text:
                    parts.append(child_text)
            value = ' '.join(parts)
            if value:
                depth = len(node.find_parents(['ul', 'ol'])) - 1
                blocks.append({'type': 'li', 'text': _clean_inline(value), 'depth': max(depth, 0)})
        elif node.name == 'table':
            source_rows = []
            for tr in node.find_all('tr'):
                cells = [
                    SourceCell(
                        text=cell.get_text(' ', strip=True),
                        row_span=max(1, int(cell.get('rowspan', 1) or 1)),
                        col_span=max(1, int(cell.get('colspan', 1) or 1)),
                    )
                    for cell in tr.find_all(['th', 'td'])
                ]
                if cells:
                    source_rows.append(cells)
            grid, merged_cells = expand_spanned_rows(source_rows)
            header, rows = block_rows_from_grid(grid)
            if header:
                block = (
                    {'type': 'table', 'header': header, 'rows': rows, 'table_source': 'html', 'merged_cells': merged_cells}
                    if merged_cells
                    else {'type': 'table', 'header': header, 'rows': rows, 'table_source': 'html'}
                )
                blocks.append(block)
    return blocks


def parse_csv_file(path):
    text = read_text_file(path)
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample)
    except csv.Error:
        dialect = csv.excel
    rows = list(csv.reader(text.splitlines(), dialect))
    rows = [[_clean_inline(cell.strip()) for cell in row] for row in rows if any(cell.strip() for cell in row)]
    if not rows:
        return []
    return [{'type': 'table', 'header': rows[0], 'rows': rows[1:], 'table_source': 'csv'}]


def _trim_grid(grid):
    while grid and not any(grid[-1]):
        grid.pop()
    max_cols = max((len(row) for row in grid), default=0)
    while max_cols > 0 and all((len(row) <= max_cols - 1 or row[max_cols - 1] == '') for row in grid):
        max_cols -= 1
    return [row[:max_cols] for row in grid]


def _xlsx_table_block(ws):
    top_left_spans = {}
    covered = set()
    for merged_range in ws.merged_cells.ranges:
        row_span = merged_range.max_row - merged_range.min_row + 1
        col_span = merged_range.max_col - merged_range.min_col + 1
        top_left_spans[(merged_range.min_row, merged_range.min_col)] = (row_span, col_span)
        for row_index in range(merged_range.min_row, merged_range.max_row + 1):
            for col_index in range(merged_range.min_col, merged_range.max_col + 1):
                if row_index == merged_range.min_row and col_index == merged_range.min_col:
                    continue
                covered.add((row_index, col_index))
    source_rows = []
    for row_index in range(1, ws.max_row + 1):
        source_row = []
        for col_index in range(1, ws.max_column + 1):
            if (row_index, col_index) in covered:
                continue
            row_span, col_span = top_left_spans.get((row_index, col_index), (1, 1))
            value = ws.cell(row=row_index, column=col_index).value
            source_row.append(SourceCell('' if value is None else _clean_inline(str(value)), row_span, col_span))
        if source_row:
            source_rows.append(source_row)
    grid, merged_cells = expand_spanned_rows(source_rows)
    header, rows = block_rows_from_grid(_trim_grid(grid))
    if not header:
        return None
    block = {
        'type': 'table',
        'header': header,
        'rows': rows,
        'table_source': 'xlsx',
        'worksheet_title': ws.title,
    }
    if merged_cells:
        block['merged_cells'] = merged_cells
    return block


def parse_xlsx(path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError('XLSX 변환에는 openpyxl이 필요함: pip install openpyxl') from exc

    wb = load_workbook(path, data_only=True)
    try:
        blocks = []
        for ws in wb.worksheets:
            table_block = _xlsx_table_block(ws)
            if table_block is None:
                continue
            blocks.append({'type': 'h', 'level': 2, 'text': ws.title})
            blocks.append(table_block)
        return blocks
    finally:
        wb.close()


def _odl_cell_text(cell):
    """tableCell.kids에서 텍스트를 추출하여 하나의 문자열로 반환."""
    parts = _odl_collect_text_parts(cell)
    return " ".join(parts)


def _odl_collect_text_parts(element):
    parts = []
    content = element.get("content", "").strip()
    if content:
        parts.append(content)
    for kid in element.get("kids", []):
        parts.extend(_odl_collect_text_parts(kid))
    for item in element.get("list items", []):
        item_content = item.get("content", "").strip()
        if item_content:
            parts.append(item_content)
        for kid in item.get("kids", []):
            parts.extend(_odl_collect_text_parts(kid))
    return parts


def _odl_span(cell, keys):
    for key in keys:
        value = cell.get(key)
        if value is None:
            continue
        try:
            return max(1, int(value))
        except (TypeError, ValueError):
            return 1
    return 1


def _odl_table_grid(element):
    """ODL 셀 좌표(row/column number, span)로 원본 표 격자를 복원.

    병합 셀(rowspan/colspan)은 앵커 위치에만 내용을 두고 병합 범위는
    빈 칸으로 유지해 열 밀림을 방지한다. 좌표 정보가 없는 구버전 JSON은
    기존 순서 기반 배열로 폴백.

    Returns:
        (grid, merges) — merges는 (row, col, row_span, col_span) 목록.
        행·열 인덱스는 헤더 포함 0부터 시작하며, 빈 행 제거 후 좌표로 보정됨.
    """
    rows = element.get("rows", [])
    flat = [cell for row in rows for cell in (row.get("cells") or [])]
    if not flat:
        return [], []
    has_coords = all(
        isinstance(cell.get("row number"), int) and isinstance(cell.get("column number"), int)
        for cell in flat
    )
    if not has_coords:
        source_rows = []
        for row in rows:
            source_row = [
                SourceCell(
                    _odl_cell_text(cell),
                    _odl_span(cell, ("rowspan", "rowSpan", "row span", "row_span")),
                    _odl_span(cell, ("colspan", "colSpan", "column span", "col span", "col_span")),
                )
                for cell in (row.get("cells") or [])
            ]
            if any(cell.text for cell in source_row):
                source_rows.append(source_row)
        grid, merged_cells = expand_spanned_rows(source_rows)
        return grid, [tuple(span) for span in merged_cells]
    n_rows = max(cell["row number"] for cell in flat)
    n_cols = max(cell["column number"] + cell.get("column span", 1) - 1 for cell in flat)
    grid = [[''] * n_cols for _ in range(n_rows)]
    merges = []
    for cell in flat:
        r = cell["row number"] - 1
        c = cell["column number"] - 1
        grid[r][c] = _odl_cell_text(cell)
        row_span = max(int(cell.get("row span", 1) or 1), 1)
        col_span = max(int(cell.get("column span", 1) or 1), 1)
        if row_span > 1 or col_span > 1:
            merges.append((r, c, row_span, col_span))

    # 빈 행 제거 + 병합 좌표 보정
    kept = [ri for ri, row in enumerate(grid) if any(cell.strip() for cell in row)]
    if len(kept) == len(grid):
        return grid, merges
    new_index = {old: new for new, old in enumerate(kept)}
    remapped = []
    for r, c, row_span, col_span in merges:
        if r not in new_index:
            continue
        span_kept = sum(1 for ri in range(r, r + row_span) if ri in new_index)
        remapped.append((new_index[r], c, max(span_kept, 1), col_span))
    return [grid[ri] for ri in kept], remapped


def _odl_text_to_block(content, default_type="p"):
    """ODL 텍스트를 항목기호 정규화를 거쳐 li 또는 p 블록으로 변환.

    '1.다음 …'처럼 점 뒤 공백이 없는 PDF 추출 텍스트도
    공문서 항목기호 규칙('1. ' 1타)으로 정규화된다.
    """
    li_result = _detect_list_item(content)
    if li_result and li_result['marker'] != '•':
        return {
            'type': 'li',
            'text': li_result['text'],
            'depth': li_result['depth'],
            'marker': li_result['marker'],
            'content': li_result['content'],
        }
    return {'type': default_type, 'text': content, **({'depth': 1} if default_type == 'li' else {})}


def _odl_element_to_blocks(element):
    """opendataloader-pdf JSON element 하나를 내부 blocks 리스트로 변환."""
    blocks = []
    etype = element.get("type", "")

    if etype == "heading":
        level = min(max(int(element.get("heading level", 1)), 1), 3)
        content = element.get("content", "").strip()
        if content:
            blocks.append({"type": "h", "level": level, "text": content})

    elif etype in ("paragraph", "caption"):
        content = element.get("content", "").strip()
        if content:
            blocks.append(_odl_text_to_block(content))

    elif etype == "table":
        grid, merges = _odl_table_grid(element)
        if grid:
            block = {"type": "table", "header": grid[0], "rows": grid[1:], "table_source": "pdf"}
            if merges:
                block["merged_cells"] = [list(span) for span in merges]
            blocks.append(block)

    elif etype == "list":
        for item in element.get("list items", []):
            content = item.get("content", "").strip()
            if content:
                blocks.append(_odl_text_to_block(content, default_type="li"))
            for child in item.get("kids", []):
                blocks.extend(_odl_element_to_blocks(child))

    elif etype == "text block":
        for child in element.get("kids", []):
            blocks.extend(_odl_element_to_blocks(child))

    # image / header / footer → 무시
    return blocks


def _odl_element_meta(element):
    """요소의 (page, x0, y0, x1, y1). 좌표 정보가 없으면 None."""
    bb = element.get('bounding box')
    page = element.get('page number')
    if not bb or len(bb) < 4 or not isinstance(page, int):
        return None
    try:
        return (page, float(bb[0]), float(bb[1]), float(bb[2]), float(bb[3]))
    except (TypeError, ValueError):
        return None


def _page_column_classifier(metas):
    """페이지 내 요소 좌표로 (열 판별 함수, 2단 여부)를 반환.

    열 판별 함수는 (x0, x1)을 받아 'left'/'right'/'full'을 돌려준다.
    """
    x_min = min(m[1] for m in metas)
    x_max = max(m[3] for m in metas)
    mid = (x_min + x_max) / 2
    tol = (x_max - x_min) * 0.05

    def column_of(x0, x1):
        if x0 < mid - tol and x1 > mid + tol:
            return 'full'
        return 'left' if (x0 + x1) / 2 < mid else 'right'

    columns = [column_of(m[1], m[3]) for m in metas]
    is_two_col = columns.count('left') >= 3 and columns.count('right') >= 3
    return column_of, is_two_col


def _odl_reading_order(elements):
    """2단 조판 페이지의 요소를 좌표 기반(좌열 전체 → 우열 전체)으로 재정렬.

    ODL이 좌/우 열 본문을 y좌표 순으로 교차 배열하는 경우(시험지 등)를
    교정한다. 중앙을 가로지르는 요소는 밴드 경계로 취급하고, 단일 컬럼
    페이지나 좌표 없는 요소가 있으면 원본 순서를 유지한다.
    """
    metas = [_odl_element_meta(e) for e in elements]
    if not elements or any(m is None for m in metas):
        return list(elements)
    pages = {}
    for element, meta in zip(elements, metas):
        pages.setdefault(meta[0], []).append((element, meta))
    ordered = []
    for page in sorted(pages):
        items = pages[page]
        classify, is_two_col = _page_column_classifier([m for _, m in items])

        def column_of(meta):
            return classify(meta[1], meta[3])

        if not is_two_col:
            ordered.extend(e for e, _ in items)
            continue
        # PDF 좌표(원점 좌하단): 위→아래 = y1 내림차순
        seq = sorted(
            ((item, column_of(item[1])) for item in items),
            key=lambda it: -it[0][1][4],
        )
        band = {'left': [], 'right': []}

        def flush_band():
            ordered.extend(e for e in band['left'])
            ordered.extend(e for e in band['right'])
            band['left'], band['right'] = [], []

        for (element, meta), col in seq:
            if col == 'full':
                flush_band()
                ordered.append(element)
            else:
                band[col].append(element)
        flush_band()
    return ordered


# ─── 변환 노트 (보완·경고 메시지를 CLI·GUI 양쪽에 전달) ─────────────────────────

_conversion_notes = []


def _add_conversion_note(message):
    _conversion_notes.append(message)
    print(message)


def pop_conversion_notes():
    notes = list(_conversion_notes)
    _conversion_notes.clear()
    return notes


# ─── ODL 누락 줄 자동 보완 (pdfplumber 교차 검증) ──────────────────────────────

_MATCH_TRANSLATE = str.maketrans({
    'ㆍ': '·', '∼': '~', '−': '-', '–': '-', '—': '-',
    ' ': '', '“': '"', '”': '"', '‘': "'", '’': "'",
})


def _norm_match_text(text):
    """엔진 간 대조용 정규화: 공백 제거 + 유사 문자 통일."""
    return re.sub(r'\s+', '', str(text or '')).translate(_MATCH_TRANSLATE)


def _odl_subtree_texts(element):
    """요소 서브트리의 모든 content 문자열 수집 (표 셀·목록 항목 포함)."""
    texts = []

    def walk(node):
        if isinstance(node, dict):
            content = str(node.get('content', '') or '')
            if content.strip():
                texts.append(content)
            for value in node.values():
                walk(value)
        elif isinstance(node, list):
            for item in node:
                walk(item)

    walk(element)
    return texts


def _group_words_into_lines(page_no, page_height, words):
    lines = []
    current, current_top = [], None
    for word in sorted(words, key=lambda w: (_word_number(w, 'top'), _word_number(w, 'x0'))):
        top = _word_number(word, 'top')
        if current_top is None or abs(top - current_top) <= 4:
            current.append(word)
            current_top = top if current_top is None else current_top
            continue
        lines.append(_line_from_words(page_no, page_height, current))
        current, current_top = [word], top
    if current:
        lines.append(_line_from_words(page_no, page_height, current))
    return lines


def _pdfplumber_page_lines(path):
    """pdfplumber로 페이지별 텍스트 줄을 좌표(PDF 좌표계)와 함께 추출.

    - 표 영역 안의 줄 제외 (셀 줄바꿈 차이로 인한 오탐 방지)
    - 2단 페이지는 열별로 줄을 구성 (좌·우 열 합성 줄 방지)
    """
    import pdfplumber

    lines = []
    with pdfplumber.open(str(path)) as pdf:
        for page_no, page in enumerate(pdf.pages, start=1):
            height = float(page.height or 0)
            table_bboxes = [t.bbox for t in (page.find_tables() or [])]
            words = [
                w for w in (page.extract_words() or [])
                if not any(_word_inside_bbox(w, bbox) for bbox in table_bboxes)
            ]
            mid = _pdfplumber_column_mid(getattr(page, 'width', 0), words)
            if mid is None:
                word_groups = [words]
            else:
                word_groups = [
                    [w for w in words if _word_center_x(w) < mid],
                    [w for w in words if _word_center_x(w) >= mid],
                ]
            for group in word_groups:
                lines.extend(_group_words_into_lines(page_no, height, group))
    return [ln for ln in lines if ln]


def _line_from_words(page_no, page_height, words):
    text = ' '.join(str(w.get('text', '')).strip() for w in words).strip()
    if not text:
        return None
    return {
        'page': page_no,
        'x0': min(_word_number(w, 'x0') for w in words),
        'x1': max(_word_number(w, 'x1') for w in words),
        # PDF 좌표계(원점 좌하단)로 변환: ODL bbox와 동일 기준
        'y_top': page_height - min(_word_number(w, 'top') for w in words),
        'text': text,
    }


_PAGE_NUMBER_LINE = re.compile(r'^[-‒–—]?\s*\d{1,3}\s*[-‒–—]?$')


def _merge_missing_lines(elements, lines):
    """ODL 결과에 없는 pdfplumber 줄을 좌표 기반 위치에 삽입.

    Returns:
        (elements, recovered_texts, warning_texts)
        - 확실한 위치를 찾은 줄은 paragraph 요소로 삽입
        - 페이지에 기준 요소가 없는 줄은 warning으로만 보고
        - 짧은 줄·쪽번호·반복 머리글(3개 페이지 이상 동일)은 무시
    """
    elements = list(elements)
    metas = [_odl_element_meta(e) for e in elements]
    if not elements or any(m is None for m in metas):
        return elements, [], []

    # 페이지별 코퍼스 (동일 텍스트가 여러 문항에 반복되는 경우 대비)
    page_corpus = {}
    for element, meta in zip(elements, metas):
        page_corpus.setdefault(meta[0], []).append(
            _norm_match_text(' '.join(_odl_subtree_texts(element)))
        )

    # ODL이 이미지로 처리한 영역(그림 내부 라벨)은 보완 대상에서 제외
    image_boxes = {}
    for element, meta in zip(elements, metas):
        if element.get('type') == 'image' and meta:
            image_boxes.setdefault(meta[0], []).append(meta[1:])

    def _inside_image(line):
        cx = (line['x0'] + line['x1']) / 2
        cy = line['y_top'] - 5
        for x0, y0, x1, y1 in image_boxes.get(line['page'], ()):
            if x0 <= cx <= x1 and y0 <= cy <= y1:
                return True
        return False

    # 머리글·꼬리말 제외: ODL 본문 요소 영역(envelope) 밖 y좌표의 줄
    page_envelope = {}
    for meta in metas:
        lo, hi = page_envelope.get(meta[0], (float('inf'), float('-inf')))
        page_envelope[meta[0]] = (min(lo, meta[2]), max(hi, meta[4]))

    def _outside_body(line):
        lo, hi = page_envelope.get(line['page'], (None, None))
        if lo is None:
            return False
        return line['y_top'] > hi + 5 or line['y_top'] - 10 < lo - 5

    def _missing_instances(page, norm, instances):
        """같은 페이지에서 norm이 등장하는 줄 인스턴스 중 ODL에 없는 것을 선별.

        동일 텍스트가 여러 문항에 반복될 수 있으므로, norm을 포함한 ODL
        요소의 y좌표와 인스턴스를 짝지어 남는 인스턴스만 누락으로 본다.
        """
        joined = ''.join(page_corpus.get(page, []))
        if norm not in joined:
            return list(instances)
        element_ys = [
            m[4] for element, m in zip(elements, metas)
            if m[0] == page and norm in _norm_match_text(' '.join(_odl_subtree_texts(element)))
        ]
        budget = max(1, len(element_ys))
        if len(instances) <= budget:
            return []
        # 요소 y와 가까운 인스턴스부터 '존재'로 매칭, 남는 것이 누락
        remaining = list(instances)
        for ey in element_ys:
            if not remaining:
                break
            closest = min(remaining, key=lambda ln: abs(ln['y_top'] - ey))
            remaining.remove(closest)
        # element_ys가 비었지만 joined에 있는 경우(조각 분할)는 1개를 존재로 간주
        if not element_ys and remaining:
            remaining = sorted(remaining, key=lambda ln: -ln['y_top'])[1:]
        return remaining

    page_joined = {page: ''.join(parts) for page, parts in page_corpus.items()}

    # (페이지, 정규화 텍스트) 단위로 묶어 누락 인스턴스 선별
    grouped = {}
    for line in lines:
        norm = _norm_match_text(line['text'])
        if len(norm) < 6 or _PAGE_NUMBER_LINE.match(line['text'].strip()):
            continue
        if _outside_body(line):
            continue
        if _inside_image(line):
            continue
        joined = page_joined.get(line['page'], '')
        if norm not in joined:
            # 부분 일치: 줄바꿈 병합 차이·그림 라벨 합성으로 인한 유사 중복 스킵
            trim = max(6, len(norm) // 3)
            if len(norm) > trim and (norm[trim:] in joined or norm[:-trim] in joined):
                continue
        grouped.setdefault((line['page'], norm), []).append(line)

    missing_lines = []
    for (page, norm), instances in grouped.items():
        missing_lines.extend(_missing_instances(page, norm, instances))

    recovered, warnings = [], []
    for line in sorted(missing_lines, key=lambda ln: (ln['page'], -ln['y_top'])):
        page_items = [(i, m) for i, m in enumerate(metas) if m and m[0] == line['page']]
        if not page_items:
            warnings.append(line['text'])
            continue
        classify, is_two_col = _page_column_classifier([m for _, m in page_items])
        candidates = page_items
        if is_two_col:
            line_col = classify(line['x0'], line['x1'])
            if line_col in ('left', 'right'):
                same_col = [
                    (i, m) for i, m in page_items
                    if classify(m[1], m[3]) == line_col
                ]
                if same_col:
                    candidates = same_col
        insert_at = None
        for i, m in candidates:
            if m[4] < line['y_top']:  # 요소 상단이 줄보다 아래 → 줄을 그 앞에
                insert_at = i
                break
        if insert_at is None:
            insert_at = candidates[-1][0] + 1
        synthetic = {
            'type': 'paragraph',
            'page number': line['page'],
            'bounding box': [line['x0'], line['y_top'] - 10, line['x1'], line['y_top']],
            'content': line['text'],
            'kids': [],
        }
        elements.insert(insert_at, synthetic)
        metas.insert(insert_at, _odl_element_meta(synthetic))
        recovered.append(line['text'])
    return elements, recovered, warnings


def _odl_data_to_blocks(data, source_path=None):
    """opendataloader-pdf JSON 최상위 객체 전체를 blocks 리스트로 변환.

    source_path가 주어지면 pdfplumber 교차 검증으로 ODL이 유실한 줄을
    자동 보완하고, 보완 불가 줄은 변환 노트로 보고한다.
    """
    elements = _odl_reading_order(data.get("kids", []))
    if source_path is not None:
        try:
            lines = _pdfplumber_page_lines(source_path)
        except Exception:
            lines = []
        if lines:
            elements, recovered, warnings = _merge_missing_lines(elements, lines)
            if recovered:
                _add_conversion_note(f'[보완] PDF 교차 검증: 누락 줄 {len(recovered)}건 자동 복구')
            for warning in warnings:
                _add_conversion_note(f'[확인 필요] 위치를 정하지 못한 누락 줄: {warning[:50]}')
    blocks = []
    for element in elements:
        blocks.extend(_odl_element_to_blocks(element))
    return blocks


def extract_pdf_blocks_odl(path):
    """opendataloader-pdf로 PDF를 JSON 추출하여 구조화된 blocks를 반환.

    Java 또는 opendataloader_pdf 패키지가 없으면 RuntimeError를 발생시킨다.
    """
    try:
        import opendataloader_pdf
    except ImportError as exc:
        raise RuntimeError(f"opendataloader_pdf 패키지 없음: {exc}") from exc

    with tempfile.TemporaryDirectory() as tmpdir:
        try:
            opendataloader_pdf.convert(
                input_path=[str(path)],
                output_dir=tmpdir,
                format="json",
            )
        except Exception as exc:
            raise RuntimeError(f"opendataloader-pdf 변환 실패: {exc}") from exc

        json_files = sorted(Path(tmpdir).glob("*.json"))
        if not json_files:
            raise RuntimeError("opendataloader-pdf: JSON 출력 파일을 찾을 수 없음")

        data = json.loads(json_files[0].read_text(encoding="utf-8"))

    blocks = _odl_data_to_blocks(data, source_path=path)
    if not blocks:
        raise RuntimeError("opendataloader-pdf: 추출된 blocks가 없음")
    return blocks


def _pdfplumber_clean_cell(value):
    if value is None:
        return ''
    text = re.sub(r'\s+', ' ', str(value).replace('\n', ' '))
    return _clean_inline(text)


def _pdfplumber_table_to_block(table_rows):
    rows = []
    for raw_row in table_rows or []:
        row = [_pdfplumber_clean_cell(cell) for cell in (raw_row or [])]
        if any(row):
            rows.append(row)
    if not rows:
        return None
    header, body = _normalize_parsed_table(rows[0], rows[1:])
    return {'type': 'table', 'header': header, 'rows': body, 'table_source': 'pdf'}


def _word_number(word, key):
    try:
        return float(word.get(key, 0) or 0)
    except (TypeError, ValueError):
        return 0.0


def _word_center_y(word):
    return (_word_number(word, 'top') + _word_number(word, 'bottom')) / 2


def _word_inside_bbox(word, bbox):
    x = (_word_number(word, 'x0') + _word_number(word, 'x1')) / 2
    y = _word_center_y(word)
    left, top, right, bottom = bbox
    return left <= x <= right and top <= y <= bottom


def _words_in_vertical_band(words, top, bottom, table_bboxes):
    band_words = []
    for word in words:
        y = _word_center_y(word)
        if y < top or y >= bottom:
            continue
        if any(_word_inside_bbox(word, bbox) for bbox in table_bboxes):
            continue
        band_words.append(word)
    return band_words


def _pdfplumber_words_to_text(words):
    lines = []
    current_line = []
    current_top = None
    for word in sorted(words, key=lambda item: (_word_number(item, 'top'), _word_number(item, 'x0'))):
        text = str(word.get('text', '')).strip()
        if not text:
            continue
        top = _word_number(word, 'top')
        if current_top is None or abs(top - current_top) <= 4:
            current_line.append(text)
            current_top = top if current_top is None else current_top
            continue
        lines.append(' '.join(current_line))
        current_line = [text]
        current_top = top
    if current_line:
        lines.append(' '.join(current_line))
    return '\n'.join(lines)


def _word_center_x(word):
    return (_word_number(word, 'x0') + _word_number(word, 'x1')) / 2


def _pdfplumber_column_mid(page_width, words):
    """2단(좌/우) 레이아웃 감지. 2단이면 분할 x좌표, 아니면 None.

    페이지 중앙을 가로지르는 단어가 적고 좌·우에 내용이 고르게 분포할 때만
    2단으로 판단한다 (시험지·자료집의 2단 조판).
    """
    try:
        page_width = float(page_width or 0)
    except (TypeError, ValueError):
        return None
    if page_width <= 0 or not words or len(words) < 40:
        return None
    mid = page_width / 2
    crossing = sum(
        1 for w in words
        if _word_number(w, 'x0') < mid - 8 and _word_number(w, 'x1') > mid + 8
    )
    if crossing / len(words) > 0.08:
        return None
    left = sum(1 for w in words if _word_center_x(w) < mid)
    right = len(words) - left
    if min(left, right) / len(words) < 0.25:
        return None
    return mid


def _pdfplumber_region_blocks(words, table_items):
    """단어·표 목록 하나의 영역(열)을 위→아래 순서로 blocks로 변환."""
    table_items = sorted(table_items, key=lambda item: item[0][1])
    table_bboxes = [bbox for bbox, _ in table_items]
    blocks = []
    cursor = 0.0
    for bbox, block in table_items:
        text = _pdfplumber_words_to_text(_words_in_vertical_band(words, cursor, bbox[1], table_bboxes))
        if text.strip():
            blocks.extend(parse_plain_text(text))
        blocks.append(block)
        cursor = max(cursor, bbox[3])
    text = _pdfplumber_words_to_text(_words_in_vertical_band(words, cursor, float('inf'), table_bboxes))
    if text.strip():
        blocks.extend(parse_plain_text(text))
    return blocks


def _pdfplumber_page_to_blocks(page):
    table_items = []
    for table in page.find_tables() or []:
        block = _pdfplumber_table_to_block(table.extract())
        if block:
            table_items.append((table.bbox, block))

    words = page.extract_words() or []
    mid = _pdfplumber_column_mid(getattr(page, 'width', 0), words)

    if mid is not None:
        # 2단 조판: 좌측 열 전체 → 우측 열 전체 순서로 읽는다
        blocks = []
        for is_left in (True, False):
            side_words = [w for w in words if (_word_center_x(w) < mid) == is_left]
            side_tables = [
                (bbox, block) for bbox, block in table_items
                if (((bbox[0] + bbox[2]) / 2) < mid) == is_left
            ]
            blocks.extend(_pdfplumber_region_blocks(side_words, side_tables))
        return blocks

    if not table_items:
        page_text = page.extract_text() or ''
        if page_text.strip():
            return parse_markdown(page_text)

    return _pdfplumber_region_blocks(words, table_items)


def extract_pdf_blocks_pdfplumber(path):
    try:
        import pdfplumber
    except ImportError as exc:
        raise RuntimeError(f'pdfplumber 없음: {exc}') from exc

    blocks = []
    try:
        with pdfplumber.open(str(path)) as pdf:
            for page in pdf.pages:
                blocks.extend(_pdfplumber_page_to_blocks(page))
    except Exception as exc:
        raise RuntimeError(f'pdfplumber 구조 추출 실패: {exc}') from exc
    if not blocks:
        raise RuntimeError('pdfplumber: 추출된 blocks가 없음')
    return blocks


def try_kordoc_pdf_text(path, kordoc_home=None):
    kordoc_dir = resolve_kordoc_dir(kordoc_home)
    if kordoc_dir is None:
        return None
    for cmd in _kordoc_commands(kordoc_dir, path):
        try:
            result = subprocess.run(
                cmd,
                cwd=str(kordoc_dir),
                capture_output=True,
                text=True,
                encoding='utf-8',
                errors='replace',
                timeout=120,
                check=False,
                env=_utf8_subprocess_env(),
            )
        except (OSError, subprocess.TimeoutExpired):
            continue
        output = (result.stdout or '').strip()
        if result.returncode == 0 and output:
            return output
    return None


def extract_pdf_text_fallback(path):
    errors = []

    try:
        import pdfplumber
        parts = []
        with pdfplumber.open(str(path)) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ''
                if text.strip():
                    parts.append(text)
        if parts:
            return '\n\n'.join(parts)
    except ImportError as exc:
        errors.append(f'pdfplumber 없음: {exc}')
    except Exception as exc:
        errors.append(f'pdfplumber 실패: {exc}')

    try:
        import fitz
        parts = []
        with fitz.open(str(path)) as doc:
            for page in doc:
                text = page.get_text('text') or ''
                if text.strip():
                    parts.append(text)
        if parts:
            return '\n\n'.join(parts)
    except ImportError as exc:
        errors.append(f'PyMuPDF 없음: {exc}')
    except Exception as exc:
        errors.append(f'PyMuPDF 실패: {exc}')

    try:
        from pypdf import PdfReader
    except ImportError as exc:
        errors.append(f'pypdf 없음: {exc}')
    else:
        try:
            reader = PdfReader(str(path))
            parts = []
            for page in reader.pages:
                text = page.extract_text() or ''
                if text.strip():
                    parts.append(text)
            if parts:
                return '\n\n'.join(parts)
        except Exception as exc:
            errors.append(f'pypdf 실패: {exc}')

    detail = '; '.join(errors) if errors else '사용 가능한 PDF 텍스트 추출기가 없음'
    raise RuntimeError(f'PDF 텍스트 추출 실패: {detail}')


def parse_pdf(path, kordoc_home=None):
    try:
        return extract_pdf_blocks_odl(path)
    except Exception as odl_exc:
        _odl_warn = str(odl_exc)

    try:
        return extract_pdf_blocks_pdfplumber(path)
    except RuntimeError as plumber_exc:
        _plumber_warn = str(plumber_exc)

    text = try_kordoc_pdf_text(path, kordoc_home=kordoc_home)

    if text is None:
        try:
            text = extract_pdf_text_fallback(path)
        except RuntimeError as fb_exc:
            raise RuntimeError(
                f'PDF 텍스트 추출 실패.\n'
                f'  opendataloader-pdf: {_odl_warn}\n'
                f'  pdfplumber 구조 추출: {_plumber_warn}\n'
                f'  fallback: {fb_exc}'
            ) from fb_exc

    if not text.strip():
        raise RuntimeError(f'PDF에서 텍스트를 추출하지 못함: {path}')
    return parse_markdown(text)


# ─── DOCX 파서 ─────────────────────────────────────────────────────────────────

def _iter_block_items(doc):
    from docx.oxml.ns import qn
    from docx.table import Table as DocxTable
    from docx.text.paragraph import Paragraph as DocxParagraph

    body = doc.element.body
    for child in body.iterchildren():
        if child.tag == qn('w:p'):
            yield DocxParagraph(child, doc)
        elif child.tag == qn('w:tbl'):
            yield DocxTable(child, doc)


def _para_text(para):
    from docx.oxml.ns import qn
    parts = []
    for run in para.runs:
        has_image = (
            run._r.find(qn('w:drawing')) is not None
            or run._r.find(qn('w:pict')) is not None
        )
        if not has_image:
            parts.append(run.text)
    return ''.join(parts).strip()


def _list_depth(para):
    from docx.oxml.ns import qn
    pPr = para._p.pPr
    if pPr is None:
        return -1
    numPr = pPr.find(qn('w:numPr'))
    if numPr is None:
        return -1
    ilvl = numPr.find(qn('w:ilvl'))
    if ilvl is None:
        return 0
    try:
        return int(ilvl.get(qn('w:val'), 0))
    except (TypeError, ValueError):
        return 0


def _docx_cell_text(tc, text_tag):
    return ''.join(node.text for node in tc.iter() if node.tag == text_tag and node.text).strip()


def _docx_grid_span(tc):
    grid_span = tc.tcPr.gridSpan if tc.tcPr is not None else None
    if grid_span is None or grid_span.val is None:
        return 1
    return max(1, int(grid_span.val))


def _docx_vmerge_value(tc):
    vmerge = tc.tcPr.vMerge if tc.tcPr is not None else None
    return '' if vmerge is None or vmerge.val is None else str(vmerge.val)


def _docx_table_grid(table):
    from docx.oxml.ns import qn
    text_tag = qn('w:t')
    grid = []
    merged_cells = []
    active_vertical = {}
    for row_index, table_row in enumerate(table._tbl.tr_lst):
        row_values = []
        col_index = 0
        for tc in table_row.tc_lst:
            while len(row_values) <= col_index:
                row_values.append('')
            col_span = _docx_grid_span(tc)
            vmerge = _docx_vmerge_value(tc)
            if vmerge == 'continue':
                merge_index = active_vertical.get(col_index)
                if merge_index is not None:
                    merged_cells[merge_index][2] += 1
                row_values[col_index] = ''
            else:
                for covered_col in range(col_index, col_index + col_span):
                    active_vertical.pop(covered_col, None)
                row_values[col_index] = _docx_cell_text(tc, text_tag)
                if col_span > 1 or vmerge == 'restart':
                    merged_cells.append([row_index, col_index, 1, col_span])
                    merge_index = len(merged_cells) - 1
                    if vmerge == 'restart':
                        for covered_col in range(col_index, col_index + col_span):
                            active_vertical[covered_col] = merge_index
            for covered_col in range(col_index + 1, col_index + col_span):
                while len(row_values) <= covered_col:
                    row_values.append('')
                row_values[covered_col] = ''
            col_index += col_span
        if any(row_values):
            grid.append(row_values)
    max_cols = max((len(row) for row in grid), default=0)
    normalized = [row + [''] * (max_cols - len(row)) for row in grid]
    return normalized, [span for span in merged_cells if span[2] > 1 or span[3] > 1]


def parse_docx(docx_path):
    try:
        from docx import Document
        from docx.table import Table as DocxTable
    except ImportError as exc:
        raise RuntimeError('DOCX 변환에는 python-docx가 필요함: pip install python-docx') from exc

    doc = Document(docx_path)
    blocks = []

    for item in _iter_block_items(doc):
        if isinstance(item, DocxTable):
            if not item.rows:
                continue
            grid, merged_cells = _docx_table_grid(item)
            header, rows = block_rows_from_grid(grid)
            if all(not h for h in header) and not rows:
                continue
            block = {'type': 'table', 'header': header, 'rows': rows, 'table_source': 'docx'}
            if merged_cells:
                block['merged_cells'] = merged_cells
            blocks.append(block)
            continue

        para = item
        style_name = para.style.name if para.style else ''
        text = _para_text(para)

        if not text:
            continue

        heading_match = re.match(r'^(?:Heading|제목|머리말)\s*(\d+)$', style_name, re.IGNORECASE)
        if heading_match:
            level = max(1, min(int(heading_match.group(1)), 3))
            blocks.append({'type': 'h', 'level': level, 'text': text})
            continue

        depth = _list_depth(para)
        if depth >= 0:
            blocks.append({'type': 'li', 'text': text, 'depth': min(depth, 7)})
            continue

        if re.search(r'[Qq]uote|인용', style_name):
            blocks.append({'type': 'bq', 'text': text})
            continue

        if re.search(r'[Cc]ode|코드', style_name):
            blocks.append({'type': 'code', 'text': text})
            continue

        if re.match(r'^(수신|경유|제목)\s*:', text):
            colon_idx = text.index(':')
            key = text[:colon_idx].strip()
            value = text[colon_idx + 1:].strip()
            blocks.append({'type': 'official_header', 'key': key, 'value': value})
            continue

        if re.search(r'[Hh]orizontal|구분선', style_name):
            blocks.append({'type': 'hr'})
            continue

        blocks.append({'type': 'p', 'text': text})

    return blocks


# ─── 확장자 자동 감지 ──────────────────────────────────────────────────────────

def detect_and_parse(file_path, kordoc_home=None):
    path = as_path(file_path)
    require_file(path)
    _conversion_notes.clear()
    ext = path.suffix.lower()
    if ext not in SUPPORTED_EXTENSIONS:
        supported = ', '.join(sorted(SUPPORTED_EXTENSIONS))
        raise ValueError(f'지원하지 않는 형식: {ext}  (지원: {supported})')
    if ext == '.md':
        return parse_markdown(read_text_file(path))
    if ext == '.txt':
        return parse_plain_text(read_text_file(path))
    if ext == '.docx':
        return parse_docx(str(path))
    if ext in ('.html', '.htm'):
        return parse_html(read_text_file(path))
    if ext == '.csv':
        return parse_csv_file(path)
    if ext == '.xlsx':
        return parse_xlsx(path)
    if ext == '.pdf':
        return parse_pdf(path, kordoc_home=kordoc_home)
    raise ValueError(f'지원하지 않는 형식: {ext}')


# ─── HWP COM 헬퍼 ─────────────────────────────────────────────────────────────

def format_hwp_startup_error(exc):
    return (
        'HWP COM 자동화 시작 실패. Hancom Office HWP 설치, COM 등록, '
        'FilePathCheckDLL 보안 모듈 등록 상태를 확인하세요. '
        f'원인: {exc}'
    )


try:
    import pywintypes as _pywintypes
    _COM_ERROR = _pywintypes.com_error
except ImportError:
    _COM_ERROR = None


def _com_call(fn, retries=3, delay=1.0):
    """Retry a callable on transient pywintypes.com_error (Windows COM instability)."""
    if _COM_ERROR is None:
        return fn()
    for attempt in range(retries):
        try:
            return fn()
        except _COM_ERROR:
            if attempt < retries - 1:
                time.sleep(delay)
            else:
                raise


def create_hwp_object(visible=True):
    try:
        import win32com.client
    except ImportError as exc:
        raise RuntimeError('HWP COM 자동화에는 pywin32가 필요함: pip install pywin32') from exc

    try:
        hwp = win32com.client.Dispatch('HWPFrame.HwpObject')
        hwp.RegisterModule('FilePathCheckDLL', 'SecurityModule')
        hwp.XHwpWindows.Item(0).Visible = visible
        return hwp
    except Exception as exc:
        raise RuntimeError(format_hwp_startup_error(exc)) from exc


def _run_hwp_preflight_worker(visible=False):
    hwp = None
    try:
        hwp = create_hwp_object(visible=visible)
        return 'HWP COM preflight OK: HWPFrame.HwpObject 생성 및 SecurityModule 등록 성공'
    finally:
        if hwp is not None:
            try:
                hwp.Quit()
            except Exception:
                pass


def run_hwp_preflight(visible=False, timeout=45):
    cmd = [
        sys.executable,
        str(Path(__file__).resolve()),
        '--_preflight-worker',
    ]
    if visible:
        cmd.append('--_preflight-visible')
    try:
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            encoding='utf-8',
            errors='replace',
            timeout=timeout,
            check=False,
            env=_utf8_subprocess_env(),
        )
    except subprocess.TimeoutExpired as exc:
        raise RuntimeError(f'HWP COM preflight timed out after {timeout} seconds.') from exc

    output = (result.stdout or '').strip()
    error = (result.stderr or '').strip()
    if result.returncode == 0:
        return output or 'HWP COM preflight OK'
    raise RuntimeError(error or output or 'HWP COM preflight failed.')


def insert_text(hwp, text):
    hwp.HAction.GetDefault('InsertText', hwp.HParameterSet.HInsertText.HSet)
    hwp.HParameterSet.HInsertText.Text = text
    hwp.HAction.Execute('InsertText', hwp.HParameterSet.HInsertText.HSet)


def break_para(hwp):
    hwp.HAction.Run('BreakPara')


def _blank_line(hwp):
    set_para_shape(hwp, align=0)
    set_char_shape(hwp, height=1300, font='body')
    break_para(hwp)


def set_char_shape(hwp, height=1300, bold=False, italic=False, font='body'):
    face_hangul = '휴먼명조' if font == 'body' else '맑은 고딕'
    face_latin = 'Arial'
    act = hwp.CreateAction('CharShape')
    pset = act.CreateSet()
    act.GetDefault(pset)
    pset.SetItem('Height', height)
    pset.SetItem('Bold', bold)
    pset.SetItem('Italic', italic)
    pset.SetItem('FaceNameHangul', face_hangul)
    pset.SetItem('FaceNameLatin', face_latin)
    act.Execute(pset)


def set_para_shape(hwp, align=0, space_before=0, space_after=0, indent_left=0, indent_first=0):
    act = hwp.CreateAction('ParagraphShape')
    pset = act.CreateSet()
    act.GetDefault(pset)
    pset.SetItem('Align', align)
    # SpaceBefore/SpaceAfter: silently ignored by HWP COM — handled via XML post-processing if needed
    # LineSpacing: 정본 §7-3 160% — COM 적용이 불확실하여 apply_official_line_spacing(header.xml)로 강제
    # LeftMargin works with 0.5× ratio (COM LeftMargin=X → hp:case hc:left=X/2)
    # Pass indent_left×2 so that hp:case stores indent_left exactly.
    # Always set explicitly (even 0) to prevent inheritance from the previous paragraph's LeftMargin.
    pset.SetItem('LeftMargin', indent_left * 2)
    # IndentFirst has no COM equivalent — apply_list_hanging_indents handles hc:intent post-processing
    act.Execute(pset)


def _rewrite_zip_entry(zip_path, entry_name, data):
    src = os.fspath(zip_path)
    fd, tmp_name = tempfile.mkstemp(suffix='.hwpx')
    os.close(fd)
    try:
        with zipfile.ZipFile(src, 'r') as zin, zipfile.ZipFile(tmp_name, 'w') as zout:
            for item in zin.infolist():
                content = data if item.filename == entry_name else zin.read(item.filename)
                zi = zipfile.ZipInfo(item.filename, item.date_time)
                zi.comment = item.comment
                zi.extra = item.extra
                zi.internal_attr = item.internal_attr
                zi.external_attr = item.external_attr
                zi.create_system = item.create_system
                zi.compress_type = item.compress_type
                zout.writestr(zi, content)
        shutil.move(tmp_name, src)
    finally:
        if os.path.exists(tmp_name):
            os.remove(tmp_name)


def _hp_tag(name):
    return f'{{http://www.hancom.co.kr/hwpml/2011/paragraph}}{name}'


def _int_attr(element, name, default=0):
    try:
        return int(element.attrib.get(name, default) or default)
    except (TypeError, ValueError):
        return default


def _coerce_table_layout(layout):
    if isinstance(layout, dict):
        return layout.get('header') or [], layout.get('rows') or []
    return layout or [], []


def _ensure_child(parent, tag_name):
    tag = _hp_tag(tag_name)
    child = parent.find(tag)
    if child is None:
        child = ET.SubElement(parent, tag)
    return child


def _set_attrs(element, attrs):
    changed = False
    for key, value in attrs.items():
        value = str(value)
        if element.attrib.get(key) != value:
            element.set(key, value)
            changed = True
    return changed


def _apply_table_cell_margin(tbl, attrs):
    margin = _ensure_child(tbl, 'cellMargin')
    return _set_attrs(margin, attrs)


def _compact_cell_paragraphs(tc, attrs):
    changed = False
    for para in tc.findall('.//hp:p', {'hp': 'http://www.hancom.co.kr/hwpml/2011/paragraph'}):
        para_pr = _ensure_child(para, 'paraPr')
        changed = _set_attrs(para_pr, attrs) or changed
    return changed


def _iter_parent_child_pairs(root):
    for parent in root.iter():
        for child in list(parent):
            yield parent, child


def _compact_paragraph_after_table(root, tbl, attrs):
    for parent, child in _iter_parent_child_pairs(root):
        if child is not tbl:
            continue
        siblings = list(parent)
        idx = siblings.index(child)
        for next_sibling in siblings[idx + 1:]:
            if next_sibling.tag == _hp_tag('p'):
                para_pr = _ensure_child(next_sibling, 'paraPr')
                return _set_attrs(para_pr, attrs)
        return False
    return False


# hc:left in hp:case (HwpUnitChar) → hc:intent
# Section markers and bullets diverge at depths 2, 4, 6 — both sets included.
_LIST_CASE_LEFT_TO_INTENT = {
    620:  -620,  # depth=0  Ⅰ. / •
    900:  -540,  # depth=1  1~9. / •
    960:  -600,  # depth=1  wide: 10.+ or 가.(depth=1 context)
    1260: -540,  # depth=2  •
    1320: -600,  # depth=2  가.
    1620: -540,  # depth=3  1) / •
    1980: -540,  # depth=4  •
    2040: -600,  # depth=4  가)
    2400: -600,  # depth=5  (1) / •
    2760: -600,  # depth=6  •
    2880: -720,  # depth=6  (가)
    3060: -540,  # depth=7  ① / •
    3420: -540,  # depth=8  ㉮ / •
}
# hp:default stores 2× values (standard HWPUNIT, not HwpUnitChar)
_LIST_DEFAULT_LEFT_TO_INTENT = {k * 2: v * 2 for k, v in _LIST_CASE_LEFT_TO_INTENT.items()}

_NS_HH = 'http://www.hancom.co.kr/hwpml/2011/head'
_NS_HP = 'http://www.hancom.co.kr/hwpml/2011/paragraph'
_NS_HC = 'http://www.hancom.co.kr/hwpml/2011/core'


def apply_list_hanging_indents(hwpx_path):
    """Post-process header.xml: set hc:intent for list paragraph styles created by set_para_shape."""
    if not os.path.exists(hwpx_path):
        return
    header_name = 'Contents/header.xml'
    try:
        with zipfile.ZipFile(hwpx_path, 'r') as zf:
            header_xml = zf.read(header_name)
        for prefix, uri in re.findall(rb'xmlns:(\w+)="([^"]+)"', header_xml[:4096]):
            ET.register_namespace(prefix.decode(), uri.decode())
        root = ET.fromstring(header_xml)
        changed = False
        for para_pr in root.iter(f'{{{_NS_HH}}}paraPr'):
            for switch in para_pr.iter(f'{{{_NS_HP}}}switch'):
                for block in switch:
                    margin = block.find(f'{{{_NS_HH}}}margin')
                    if margin is None:
                        continue
                    left_elem = margin.find(f'{{{_NS_HC}}}left')
                    intent_elem = margin.find(f'{{{_NS_HC}}}intent')
                    if left_elem is None or intent_elem is None:
                        continue
                    left_val = int(left_elem.get('value', '0'))
                    is_case = block.tag == f'{{{_NS_HP}}}case'
                    lookup = _LIST_CASE_LEFT_TO_INTENT if is_case else _LIST_DEFAULT_LEFT_TO_INTENT
                    if left_val in lookup:
                        new_intent = str(lookup[left_val])
                        if intent_elem.get('value') != new_intent:
                            intent_elem.set('value', new_intent)
                            changed = True
        if not changed:
            return
        _rewrite_zip_entry(hwpx_path, header_name, ET.tostring(root, encoding='utf-8', xml_declaration=True))
    except Exception as e:
        print(f'[경고] 목록 내어쓰기 후처리 실패: {e}', file=sys.stderr)


# 공문서 목록 마커 패턴 (텍스트가 이 패턴으로 시작하면 목록 항목)
_LIST_MARKER_RE = re.compile(
    r'^(?:'
    r'[ⅠⅡⅢⅣⅤⅥⅦⅧⅨⅩ]+\.?[ \t]+\S'          # Ⅰ. 총칙
    r'|\d{1,2}\.(?!\d)[ \t]+\S'              # 1. 목적 (not 2026. 3.)
    r'|[가나다라마바사아자차카타파하]\.[ \t]+\S'       # 가. 방침
    r'|\d{1,2}\)[ \t]+\S'                    # 1) 세부
    r'|[가나다라마바사아자차카타파하]\)[ \t]+\S'       # 가) 대상
    r'|\(\d{1,2}\)[ \t]+\S'                  # (1) 내용
    r'|\([가나다라마바사아자차카타파하]\)[ \t]+\S'     # (가) 방법
    r'|[①②③④⑤⑥⑦⑧⑨⑩⑪⑫⑬⑭⑮⑯⑰⑱⑲⑳][ \t]*\S'  # ① 확인
    r'|[㉮㉯㉰㉱㉲㉳㉴㉵㉶㉷㉸㉹㉺㉻][ \t]*\S'      # ㉮ 보완
    r'|[•\-][ \t]+\S'                        # bullet/dash
    r')'
)


def fix_body_text_prid(hwpx_path):
    """Post-process section0.xml: fix body text paragraphs assigned to list paraPr IDs.

    HWP COM's SetItem('LeftMargin', 0) is silently ignored (treated as no-op).
    As a result, body text paragraphs that follow list items inherit the list's
    paraPr (left=900 etc.) instead of the body text paraPr (left=0 = paraPr=0).
    This function corrects those assignments in the generated HWPX.
    """
    if not os.path.exists(hwpx_path):
        return
    section_name = 'Contents/section0.xml'
    try:
        with zipfile.ZipFile(hwpx_path, 'r') as zf:
            header_xml = zf.read('Contents/header.xml')
            section_xml = zf.read(section_name)
        for source in (header_xml, section_xml):
            for prefix, uri in re.findall(rb'xmlns:(\w+)="([^"]+)"', source[:4096]):
                ET.register_namespace(prefix.decode(), uri.decode())
        hroot = ET.fromstring(header_xml)
        sroot = ET.fromstring(section_xml)

        # Collect paraPr IDs whose left value is a known list body-stop
        list_prid_set: set[str] = set()
        for para_pr in hroot.iter(f'{{{_NS_HH}}}paraPr'):
            pid = para_pr.get('id', '')
            if not pid:
                continue
            for switch in para_pr.iter(f'{{{_NS_HP}}}switch'):
                for block in switch:
                    if block.tag != f'{{{_NS_HP}}}case':
                        continue
                    margin = block.find(f'{{{_NS_HH}}}margin')
                    if margin is None:
                        continue
                    left_elem = margin.find(f'{{{_NS_HC}}}left')
                    if left_elem is None:
                        continue
                    if int(left_elem.get('value', '0')) in _LIST_CASE_LEFT_TO_INTENT:
                        list_prid_set.add(pid)

        if not list_prid_set:
            return

        # Fix paragraphs that use a list paraPr but are not list items
        changed = False
        for p in sroot.iter(f'{{{_NS_HP}}}p'):
            prid = p.get('paraPrIDRef', '')
            if prid not in list_prid_set:
                continue
            text = ''.join(t.text or '' for t in p.iter(f'{{{_NS_HP}}}t'))
            if text and _LIST_MARKER_RE.match(text):
                continue  # Proper list item — keep
            # Body text or blank paragraph: reset to Normal (paraPr=0)
            p.set('paraPrIDRef', '0')
            changed = True

        if changed:
            _rewrite_zip_entry(
                hwpx_path, section_name,
                ET.tostring(sroot, encoding='utf-8', xml_declaration=True),
            )
    except Exception as e:
        print(f'[경고] 본문 단락 paraPr 보정 실패: {e}', file=sys.stderr)


# 페이지 여백 (HWPX_작성규칙 7-2): 상 25 / 하 20 / 좌우 25 / 머리말·꼬리말 10 (mm)
_MM_TO_HWPUNIT = 7200 / 25.4
_OFFICIAL_PAGE_MARGINS = {
    'top': round(25 * _MM_TO_HWPUNIT),
    'bottom': round(20 * _MM_TO_HWPUNIT),
    'left': round(25 * _MM_TO_HWPUNIT),
    'right': round(25 * _MM_TO_HWPUNIT),
    'header': round(10 * _MM_TO_HWPUNIT),
    'footer': round(10 * _MM_TO_HWPUNIT),
}
_OFFICIAL_LINE_SPACING = 160  # 정본 §7-3: 줄 간격 160% (apply_official_line_spacing가 header.xml에 강제)


def apply_official_line_spacing(hwpx_path):
    """Post-process header.xml: 모든 단락 속성(paraPr)의 줄 간격을 160%로 강제 (정본 §7-3).

    단락 속성은 header.xml의 hh:paraPr 정의에 저장되고 hp:p가 paraPrIDRef로 참조한다.
    기존 hh:lineSpacing이 있으면 값만 갱신(자식 순서 위험 0), 없으면 switch 뒤에 삽입한다.
    """
    if not os.path.exists(hwpx_path):
        return
    header_name = 'Contents/header.xml'
    ls_tag = f'{{{_NS_HH}}}lineSpacing'
    switch_tag = f'{{{_NS_HP}}}switch'
    try:
        with zipfile.ZipFile(hwpx_path, 'r') as zf:
            header_xml = zf.read(header_name)
        for prefix, uri in re.findall(rb'xmlns:(\w+)="([^"]+)"', header_xml[:4096]):
            ET.register_namespace(prefix.decode(), uri.decode())
        root = ET.fromstring(header_xml)
        changed = False
        for para_pr in root.iter(f'{{{_NS_HH}}}paraPr'):
            line_spacing = para_pr.find(ls_tag)
            if line_spacing is None:
                line_spacing = ET.Element(ls_tag)
                children = list(para_pr)
                insert_at = len(children)
                for index, child in enumerate(children):
                    if child.tag == switch_tag:
                        insert_at = index + 1
                        break
                para_pr.insert(insert_at, line_spacing)
            if _set_attrs(line_spacing, {'type': 'PERCENT', 'value': str(_OFFICIAL_LINE_SPACING), 'unit': 'HWPUNIT'}):
                changed = True
        if changed:
            _rewrite_zip_entry(hwpx_path, header_name, ET.tostring(root, encoding='utf-8', xml_declaration=True))
    except Exception as e:
        print(f'[경고] 줄 간격 후처리 실패: {e}', file=sys.stderr)


# 정본 §7-3 제목 단락 간격 (앞, 뒤) — 1pt = 100 HWPUNIT
_HEADING_SPACING = {
    'H1': (500, 250),
    'H2': (400, 200),
    'H3': (300, 150),
}


def _charpr_level_map(header_root):
    """charPr id → 제목 레벨('H1'/'H2'/'H3') 또는 None(본문). 정본 §7-1 크기 기준."""
    level = {}
    for cp in header_root.iter(f'{{{_NS_HH}}}charPr'):
        cid = cp.get('id')
        height = int(cp.get('height') or 0)
        bold = cp.find(f'{{{_NS_HH}}}bold') is not None
        if height >= 1600:
            level[cid] = 'H1'
        elif height >= 1400:
            level[cid] = 'H2'
        elif height >= 1300 and bold:
            level[cid] = 'H3'
        else:
            level[cid] = None
    return level


def _apply_para_margin(para_pr, prev, nxt):
    margins = list(para_pr.iter(f'{{{_NS_HH}}}margin'))
    if not margins:
        return False
    changed = False
    for margin in margins:
        for tag, value in (('prev', prev), ('next', nxt)):
            elem = margin.find(f'{{{_NS_HC}}}{tag}')
            if elem is None:
                elem = ET.SubElement(margin, f'{{{_NS_HC}}}{tag}')
            if elem.get('value') != str(value) or elem.get('unit') != 'HWPUNIT':
                elem.set('value', str(value))
                elem.set('unit', 'HWPUNIT')
                changed = True
    return changed


def _set_heading_para_spacing(header_root, section_root):
    """제목 단락(H1~H3)에만 §7-3 앞/뒤 간격을 설정한다.

    제목이 본문·다른 레벨과 paraPr을 공유하면 본문에 잘못 적용되지 않도록 건너뛴다.
    반환: (changed, skipped) — skipped는 공유로 미적용된 paraPr 수.
    """
    char_level = _charpr_level_map(header_root)
    usage = {}  # paraPrIDRef → 사용된 레벨 집합(None = 본문)
    for p in section_root.iter(f'{{{_NS_HP}}}p'):
        ppid = p.get('paraPrIDRef')
        if ppid is None:
            continue
        plevel = None
        for run in p.iter(f'{{{_NS_HP}}}run'):
            lv = char_level.get(run.get('charPrIDRef'))
            if lv is not None and (plevel is None or lv < plevel):  # H1<H2<H3 우선
                plevel = lv
        usage.setdefault(ppid, set()).add(plevel)
    para_by_id = {e.get('id'): e for e in header_root.iter(f'{{{_NS_HH}}}paraPr')}
    changed = False
    skipped = 0
    for ppid, levels in usage.items():
        heading_levels = {lv for lv in levels if lv is not None}
        if not heading_levels:
            continue  # 본문 전용 paraPr → 변경 없음
        if len(levels) == 1 and len(heading_levels) == 1:
            para_pr = para_by_id.get(ppid)
            prev, nxt = _HEADING_SPACING[next(iter(heading_levels))]
            if para_pr is not None and _apply_para_margin(para_pr, prev, nxt):
                changed = True
        else:
            skipped += 1  # 제목이 본문/타 레벨과 paraPr 공유 → 안전하게 미적용
    return changed, skipped


def apply_official_paragraph_spacing(hwpx_path):
    """Post-process header.xml: 제목 단락 앞/뒤 간격 (정본 §7-3).

    COM이 무시하는 SpaceBefore/After를 paraPr margin(hc:prev/next)으로 실제 반영한다.
    본문은 0/0(무변경). 제목이 본문과 paraPr을 공유하는 경우는 건너뛰고 참고 로그를 남긴다.
    """
    if not os.path.exists(hwpx_path):
        return
    header_name = 'Contents/header.xml'
    section_name = 'Contents/section0.xml'
    try:
        with zipfile.ZipFile(hwpx_path, 'r') as zf:
            header_xml = zf.read(header_name)
            section_xml = zf.read(section_name)
        for prefix, uri in re.findall(rb'xmlns:(\w+)="([^"]+)"', header_xml[:4096]):
            ET.register_namespace(prefix.decode(), uri.decode())
        header_root = ET.fromstring(header_xml)
        section_root = ET.fromstring(section_xml)
        changed, skipped = _set_heading_para_spacing(header_root, section_root)
        if skipped:
            print(f'  [참고] 제목 단락 간격: paraPr 공유로 {skipped}건 미적용 (§7-3)', file=sys.stderr)
        if changed:
            _rewrite_zip_entry(hwpx_path, header_name, ET.tostring(header_root, encoding='utf-8', xml_declaration=True))
    except Exception as e:
        print(f'[경고] 단락 간격 후처리 실패: {e}', file=sys.stderr)


def apply_official_page_margins(hwpx_path):
    """Post-process section0.xml: 공문서 페이지 여백 적용."""
    if not os.path.exists(hwpx_path):
        return
    section_name = 'Contents/section0.xml'
    try:
        with zipfile.ZipFile(hwpx_path, 'r') as zf:
            section_xml = zf.read(section_name)
        for prefix, uri in re.findall(rb'xmlns:(\w+)="([^"]+)"', section_xml[:4096]):
            ET.register_namespace(prefix.decode(), uri.decode())
        root = ET.fromstring(section_xml)
        changed = False
        for page_pr in root.iter(f'{{{_NS_HP}}}pagePr'):
            margin = page_pr.find(f'{{{_NS_HP}}}margin')
            if margin is None:
                continue
            for key, value in _OFFICIAL_PAGE_MARGINS.items():
                if margin.get(key) != str(value):
                    margin.set(key, str(value))
                    changed = True
        if changed:
            _rewrite_zip_entry(hwpx_path, section_name, ET.tostring(root, encoding='utf-8', xml_declaration=True))
    except Exception as e:
        print(f'[경고] 페이지 여백 후처리 실패: {e}', file=sys.stderr)


def _section_text_width(root):
    """secPr의 페이지 크기·여백으로 본문 폭(HWPUNIT)을 계산. 없으면 None."""
    page_pr = root.find(f'.//{{{_NS_HP}}}pagePr')
    if page_pr is None:
        return None
    page_width = _int_attr(page_pr, 'width', 0)
    margin = page_pr.find(f'{{{_NS_HP}}}margin')
    if page_width <= 0 or margin is None:
        return None
    width = (
        page_width
        - _int_attr(margin, 'left', 0)
        - _int_attr(margin, 'right', 0)
        - _int_attr(margin, 'gutter', 0)
    )
    return width if width > 1000 else None


def apply_table_layout_profiles(hwpx_path, table_layouts):
    _apply_table_layout_profiles_new(hwpx_path, table_layouts)


def apply_table_width_profiles(hwpx_path, table_headers):
    table_layouts = [{'header': header or [], 'rows': []} for header in (table_headers or [])]
    apply_table_layout_profiles(hwpx_path, table_layouts)


def insert_table(hwp, header, rows, table_role=None, column_widths=None, table_source=None, worksheet_title=None, merged_cells=None):
    all_rows = ([header] if header else []) + rows
    if not all_rows:
        return
    num_rows = len(all_rows)
    num_cols = max(len(r) for r in all_rows)
    layout = table_layout_for(
        header or [],
        rows,
        table_role=table_role if isinstance(table_role, str) else None,
        column_widths=column_widths if isinstance(column_widths, list) else None,
        table_source=table_source if isinstance(table_source, str) else None,
        worksheet_title=worksheet_title if isinstance(worksheet_title, str) else None,
        merged_cells=merged_cells if isinstance(merged_cells, list) else None,
    )
    col_widths = layout.column_widths
    row_heights = calc_row_heights(header or [], rows, col_widths)
    act = hwp.CreateAction('TableCreate')
    pset = act.CreateSet()
    act.GetDefault(pset)
    pset.SetItem('Rows', num_rows)
    pset.SetItem('Cols', num_cols)
    pset.SetItem('WidthType', 0)
    pset.SetItem('HeightType', 0)
    pset.SetItem('AutoHeight', True)
    for key, value in (('WidthValue', sum(col_widths)), ('HeightValue', sum(row_heights))):
        try:
            pset.SetItem(key, value)
        except Exception:
            pass
    act.Execute(pset)
    width_adjust_failed = False
    moved_right = 0
    try:
        for ci, w in enumerate(col_widths):
            sel_act = hwp.CreateAction('TableColWidth')
            if sel_act is None:
                raise RuntimeError('TableColWidth action unavailable')
            sel_pset = sel_act.CreateSet()
            sel_act.GetDefault(sel_pset)
            sel_pset.SetItem('Width', w)
            sel_act.Execute(sel_pset)
            if ci < num_cols - 1:
                hwp.HAction.Run('TableRightCell')
                moved_right += 1
    except Exception as e:
        width_adjust_failed = True
        print(f'[경고] 열 너비 조정 실패: {e}')
    finally:
        for _ in range(moved_right):
            hwp.HAction.Run('TableLeftCell')
    first_cell = True
    for ri, row in enumerate(all_rows):
        is_header = (ri == 0 and header is not None)
        for ci in range(num_cols):
            if not first_cell:
                hwp.HAction.Run('TableRightCell')
            first_cell = False
            cell_text = row[ci] if ci < len(row) else ''
            if is_header:
                set_para_shape(hwp, align=layout.style.header_align)
                set_char_shape(hwp, height=1200, bold=True, font='table')
            else:
                set_para_shape(hwp, align=layout.style.body_align)
                set_char_shape(hwp, height=1200, font='table')
            if cell_text:
                insert_text(hwp, cell_text)
    hwp.HAction.Run('MoveDocEnd')
    break_para(hwp)


# ─── 문서 빌드 ─────────────────────────────────────────────────────────────────

def build_doc(hwp, blocks):
    first_depth1_li_seen = False

    for i, blk in enumerate(blocks):
        t = blk.get('type')

        if t == 'h':
            lv = blk['level']
            heights = {1: 1600, 2: 1400, 3: 1300}
            sbefore = {1: 500, 2: 400, 3: 300}
            safter = {1: 250, 2: 200, 3: 150}
            set_para_shape(hwp, align=1, space_before=sbefore.get(lv, 300), space_after=safter.get(lv, 150))
            set_char_shape(hwp, height=heights.get(lv, 1300), bold=True, font='body')
            insert_text(hwp, blk['text'])
            break_para(hwp)
            set_para_shape(hwp, align=0)
            set_char_shape(hwp, height=1300, font='body')

        elif t == 'p':
            set_para_shape(hwp, align=0)
            set_char_shape(hwp, height=1300, font='body')
            insert_text(hwp, blk['text'])
            break_para(hwp)

        elif t == 'li':
            depth = blk.get('depth', 0)
            if depth == 1:
                if first_depth1_li_seen:
                    _blank_line(hwp)
                first_depth1_li_seen = True
            para_shape = official_list_para_shape(depth, blk.get('marker', ''))
            set_para_shape(
                hwp,
                align=para_shape['align'],
                space_before=para_shape.get('space_before', 0),
                space_after=para_shape.get('space_after', 0),
                indent_left=para_shape['indent_left'],
                indent_first=para_shape['indent_first'],
            )
            # Section markers (1. 가. Ⅰ. etc.) are bold; plain bullets (•) are not
            marker = blk.get('marker', '')
            bold = depth <= 2 and marker != '•'
            set_char_shape(hwp, height=1300, bold=bold, font='body')
            insert_text(hwp, blk['text'])
            break_para(hwp)

        elif t == 'bq':
            set_para_shape(hwp, align=1, indent_left=600)
            set_char_shape(hwp, height=1200, italic=True, font='body')
            insert_text(hwp, blk['text'])
            break_para(hwp)

        elif t == 'code':
            set_para_shape(hwp, align=1, indent_left=600)
            set_char_shape(hwp, height=1100, font='table')
            insert_text(hwp, blk['text'])
            break_para(hwp)

        elif t == 'hr':
            set_para_shape(hwp, align=3)
            set_char_shape(hwp, height=1000, font='body')
            insert_text(hwp, '─' * 30)
            break_para(hwp)

        elif t == 'table':
            if i > 0:
                _blank_line(hwp)
            set_para_shape(hwp, align=0)
            set_char_shape(hwp, height=1200, font='table')
            insert_table(
                hwp,
                blk.get('header'),
                blk.get('rows', []),
                table_role=blk.get('table_role'),
                column_widths=blk.get('column_widths'),
                table_source=blk.get('table_source'),
                worksheet_title=blk.get('worksheet_title'),
                merged_cells=blk.get('merged_cells') or blk.get('merges'),
            )
            _blank_line(hwp)

        elif t == 'official_header':
            set_para_shape(hwp, align=0)
            set_char_shape(hwp, height=1300, font='body')
            insert_text(hwp, blk['key'] + ': ' + blk['value'])
            break_para(hwp)

        elif t == 'attachment':
            # 붙임 연속 항목(2.~)은 '붙임  ' 너비(시각폭 6자)만큼 들여 정렬
            indent = 720 if blk.get('cont') else 0
            set_para_shape(hwp, align=0, indent_left=indent)
            set_char_shape(hwp, height=1300, font='body')
            insert_text(hwp, blk['text'])
            break_para(hwp)


def _ends_with_end_mark(text):
    return (text or '').strip().rstrip('.').endswith('끝')


def append_end_mark_blocks(blocks):
    """공문서 '끝' 표시 규칙 (시행규칙 제4조제5항) — 블록 전처리.

    - 본문/붙임 마지막 글자 뒤 같은 줄에 2타 띄우고 '끝.'
    - 표로 끝난 경우 표 아래 왼쪽에서 1타 띄우고 '끝.' (별도 단락)
    - 이미 '끝'·'이하 빈칸'으로 마감된 문서는 그대로 둠
    """
    if not blocks:
        return blocks
    last = blocks[-1]
    if _ends_with_end_mark(last.get('text', '')):
        return blocks
    if last['type'] == 'table':
        last_rows = last.get('rows') or []
        if last_rows:
            last_row_text = ' '.join(last_rows[-1]).strip()
            if _ends_with_end_mark(last_row_text) or last_row_text == '이하 빈칸':
                return blocks
        return blocks + [{'type': 'p', 'text': ' 끝.'}]
    if last['type'] in ('p', 'li', 'bq', 'attachment'):
        updated = dict(last)
        updated['text'] = (updated.get('text', '') or '').rstrip() + '  끝.'
        return blocks[:-1] + [updated]
    return blocks + [{'type': 'p', 'text': '  끝.'}]


_OFFICIAL_DATE_PATTERN = re.compile(
    r'(?<![\d.])(\d{4})\s*\.\s*(\d{1,2})\s*\.\s*(\d{1,2})\s*\.?(?!\d)'
)


def _format_official_date(match):
    year, month, day = match.group(1), int(match.group(2)), int(match.group(3))
    if not (1 <= month <= 12 and 1 <= day <= 31):
        return match.group(0)
    return f'{year}. {month}. {day}.'


def normalize_official_dates(blocks):
    """날짜 표기 규칙 (행정업무운영규정 영 제7조) — '2026.3.22' → '2026. 3. 22.'"""
    result = []
    for blk in blocks:
        if blk.get('type') in ('p', 'li', 'bq', 'attachment', 'official_header', 'h') and blk.get('text'):
            updated = dict(blk)
            updated['text'] = _OFFICIAL_DATE_PATTERN.sub(_format_official_date, updated['text'])
            if 'value' in updated and updated.get('value'):
                updated['value'] = _OFFICIAL_DATE_PATTERN.sub(_format_official_date, updated['value'])
            result.append(updated)
        else:
            result.append(blk)
    return result


_SINO_DIGITS = '영일이삼사오육칠팔구'
_SINO_PLACE = ('', '십', '백', '천')
_SINO_GROUP = ('', '만', '억', '조', '경')


def _sino_korean_amount(value):
    """정수를 한글 금액 표기로 변환 (정본 §1-2). 예: 113560 → 일십일만삼천오백육십.

    각 자리 숫자를 모두 표기(일십·일백 형식)하는 공문 금액 표기 방식.
    """
    n = int(value)
    if n == 0:
        return '영'
    sign = '마이너스' if n < 0 else ''
    n = abs(n)
    groups = []
    while n > 0:
        groups.append(n % 10000)
        n //= 10000
    parts = []
    for index in range(len(groups) - 1, -1, -1):
        group = groups[index]
        if group == 0:
            continue
        digits = ''
        for place in range(3, -1, -1):
            digit = (group // (10 ** place)) % 10
            if digit:
                digits += _SINO_DIGITS[digit] + _SINO_PLACE[place]
        parts.append(digits + _SINO_GROUP[index])
    return sign + ''.join(parts)


_OFFICIAL_AMOUNT_PATTERN = re.compile(
    r'금?(\d{1,3}(?:,\d{3})+|\d+)\s*원(?!\s*\()'
)


def _format_official_amount(match):
    raw = match.group(1)
    digits = int(raw.replace(',', ''))
    return f'금{raw}원(금{_sino_korean_amount(digits)}원)'


def normalize_official_amounts(blocks):
    """금액 표기 규칙 (정본 §1-2) — '113,560원' → '금113,560원(금일십일만삼천오백육십원)'.

    이미 한글 병기된 금액(뒤에 '(...'가 오는 경우)은 건드리지 않는다.
    """
    result = []
    for blk in blocks:
        if blk.get('type') in ('p', 'li', 'bq', 'attachment', 'official_header', 'h') and blk.get('text'):
            updated = dict(blk)
            updated['text'] = _OFFICIAL_AMOUNT_PATTERN.sub(_format_official_amount, updated['text'])
            if 'value' in updated and updated.get('value'):
                updated['value'] = _OFFICIAL_AMOUNT_PATTERN.sub(_format_official_amount, updated['value'])
            result.append(updated)
        else:
            result.append(blk)
    return result


# ─── 공공언어 스타일 린트 (정본 §1-1, 경고 수준·비강제) ──────────────────────────

_STYLE_TEXT_TYPES = ('p', 'li', 'bq', 'attachment', 'official_header', 'h')

# (피할 표현, 권장 표현, 뒤에 오면 제외할 접미사) — 정본 §1-1 공공언어 순화 예시
_STYLE_LINT_RULES = (
    ('금일', '오늘', None),
    ('향후', '앞으로', None),
    ('만전을 기해', '최선을 다할', None),
    ('에 있어서', '에서', None),
    ('에 위치한', '에 있는', None),
    ('에 의거', '에 따름', None),
    ('실시', '함/한다', '간'),  # '실시간'은 순화 대상 아님
)


def lint_official_style(blocks):
    """정본 §1-1 공공언어 순화·병렬('및') 경고를 반환한다(비강제, 텍스트 미수정)."""
    texts = []
    for blk in blocks:
        if blk.get('type') in _STYLE_TEXT_TYPES:
            if blk.get('text'):
                texts.append(blk['text'])
            if blk.get('value'):
                texts.append(blk['value'])
    combined = '\n'.join(texts)
    notes = []
    for avoid, prefer, exclude in _STYLE_LINT_RULES:
        if exclude:
            found = re.search(re.escape(avoid) + f'(?!{re.escape(exclude)})', combined)
        else:
            found = avoid in combined
        if found:
            notes.append(f"[확인 필요] 공공언어 순화: '{avoid}' → '{prefer}' 권장 (정본 §1-1)")
    if '및' in combined:
        notes.append("[확인 필요] '및' 사용 — '와/과/·'로 병렬관계 명확화 검토 (정본 §1-1)")
    return notes


# ─── 변환 실행 ─────────────────────────────────────────────────────────────────

def build_output_path(src_path, output_dir):
    src = as_path(src_path)
    out_dir = as_path(output_dir) if output_dir else src.parent
    out_dir.mkdir(parents=True, exist_ok=True)
    candidate = out_dir / f'{src.stem}.hwpx'
    if not candidate.exists():
        return candidate
    for idx in range(2, 1000):
        candidate = out_dir / f'{src.stem} - {idx}.hwpx'
        if not candidate.exists():
            return candidate
    raise FileExistsError(f'저장 가능한 파일명을 찾지 못함: {out_dir / (src.stem + ".hwpx")}')


def _output_manifest_path(output_dir):
    return as_path(output_dir) / OUTPUT_MANIFEST_NAME


def read_output_manifest(output_dir):
    manifest_path = _output_manifest_path(output_dir)
    if not manifest_path.exists():
        return []
    try:
        data = json.loads(manifest_path.read_text(encoding='utf-8'))
    except (OSError, json.JSONDecodeError) as exc:
        raise RuntimeError(f'출력 폴더 manifest를 읽을 수 없음: {manifest_path}') from exc
    files = data.get('files', [])
    if not isinstance(files, list):
        raise RuntimeError(f'출력 폴더 manifest 형식이 올바르지 않음: {manifest_path}')
    return [str(item) for item in files if isinstance(item, str)]


def write_output_manifest(output_dir, output_files):
    out_dir = as_path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    names = []
    for output_file in output_files:
        candidate = Path(output_file)
        if not candidate.is_absolute():
            if '..' in candidate.parts:
                raise RuntimeError(f'manifest에 출력 폴더 밖의 파일을 기록할 수 없음: {output_file}')
            names.append(candidate.as_posix())
            continue
        output_path = candidate.expanduser().resolve()
        try:
            names.append(output_path.relative_to(out_dir).as_posix())
        except ValueError as exc:
            raise RuntimeError(f'manifest에 출력 폴더 밖의 파일을 기록할 수 없음: {output_path}') from exc
    payload = {
        'app': 'anyway_to_hwpx',
        'files': sorted(dict.fromkeys(names)),
    }
    _output_manifest_path(out_dir).write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding='utf-8',
    )


def prepare_output_dir(output_dir, empty_output_folder=False):
    out_dir = as_path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    if not empty_output_folder:
        return out_dir

    manifest_path = _output_manifest_path(out_dir)
    children = [child for child in out_dir.iterdir() if child.name != OUTPUT_MANIFEST_NAME]
    if children and not manifest_path.exists():
        raise RuntimeError(
            f'출력 폴더가 비어 있지 않고 이 앱의 manifest가 없음: {out_dir}\n'
            f'다른 파일 보호를 위해 --empty-output-folder를 중단함.'
        )

    managed_files = set(read_output_manifest(out_dir))
    unknown_files = []
    for child in children:
        rel_name = child.relative_to(out_dir).as_posix()
        if rel_name not in managed_files:
            unknown_files.append(rel_name)
    if unknown_files:
        preview = ', '.join(unknown_files[:5])
        raise RuntimeError(
            f'출력 폴더에 manifest가 관리하지 않는 파일이 있음: {preview}\n'
            f'다른 파일 보호를 위해 --empty-output-folder를 중단함.'
        )

    for child in children:
        if child.is_dir():
            shutil.rmtree(child)
        else:
            child.unlink()
    write_output_manifest(out_dir, [])
    return out_dir


def record_output_file(output_dir, output_file):
    if output_dir is None:
        return
    out_dir = as_path(output_dir)
    if not _output_manifest_path(out_dir).exists():
        return
    files = read_output_manifest(out_dir)
    output_path = as_path(output_file)
    try:
        rel_name = output_path.relative_to(out_dir).as_posix()
    except ValueError:
        return
    files.append(rel_name)
    write_output_manifest(out_dir, files)


def convert_file(
    hwp,
    src_path,
    hwpx_path,
    insert_end_mark=False,
    kordoc_home=None,
    diagnose_stage: DiagnoseStageReporter | None = None,
):
    src = as_path(src_path)
    out = as_path(hwpx_path)
    if diagnose_stage:
        diagnose_stage('parse_source')
    blocks = detect_and_parse(src, kordoc_home=kordoc_home)
    notes = pop_conversion_notes()
    if insert_end_mark:
        blocks = normalize_official_dates(blocks)
        blocks = normalize_official_amounts(blocks)
        notes.extend(lint_official_style(blocks))  # 정본 §1-1 순화 경고(비강제)
        blocks = append_end_mark_blocks(blocks)
    table_layouts = [
        {
            'header': blk.get('header') or [],
            'rows': blk.get('rows') or [],
            'table_role': blk.get('table_role') or '',
            'column_widths': blk.get('column_widths') or [],
            'table_source': blk.get('table_source') or '',
            'worksheet_title': blk.get('worksheet_title') or '',
            'merged_cells': blk.get('merged_cells') or blk.get('merges') or [],
        }
        for blk in blocks
        if blk.get('type') == 'table'
    ]

    if diagnose_stage:
        diagnose_stage('XHwpDocuments.Add')
    _com_call(lambda: hwp.XHwpDocuments.Add(isTab=False))
    time.sleep(0.5)
    doc = hwp.XHwpDocuments.Item(hwp.XHwpDocuments.Count - 1)

    try:
        if diagnose_stage:
            diagnose_stage('build_doc')
        build_doc(hwp, blocks)
        if diagnose_stage:
            diagnose_stage('SaveAs')
        _com_call(lambda: hwp.SaveAs(str(out), 'HWPX', 'lock:false'))
        time.sleep(0.5)
    finally:
        if diagnose_stage:
            diagnose_stage('doc.Close')
        _com_call(lambda: doc.Close(isDirty=False))
        time.sleep(0.3)
    if diagnose_stage:
        diagnose_stage('postprocess')
    apply_official_page_margins(out)  # 여백 먼저 — 표 폭이 본문 폭 기준으로 계산되도록
    apply_table_layout_profiles(out, table_layouts)
    apply_list_hanging_indents(out)
    fix_body_text_prid(out)
    apply_official_line_spacing(out)  # 정본 §7-3: 줄 간격 160% (header.xml paraPr)
    apply_official_paragraph_spacing(out)  # 정본 §7-3: 제목 단락 앞/뒤 간격 (header.xml paraPr)
    if diagnose_stage:
        diagnose_stage('finalize')
    ext = src.suffix.upper().lstrip('.')
    print(f'[완료] {ext} → {out.name}')
    return {'notes': notes}


def main(argv=None):
    parser = argparse.ArgumentParser(
        description='Markdown / TXT / DOCX / HTML / CSV / XLSX / PDF → HWPX 변환 (HWP COM 방식)',
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument('files', nargs='*', help='변환할 파일 경로')
    parser.add_argument('-o', '--output-dir', default=None, help='저장할 폴더 경로 (기본: 입력 파일과 같은 폴더)')
    parser.add_argument('--empty-output-folder', action='store_true', help='변환 전 앱 manifest가 관리하는 출력 폴더 파일만 비움')
    parser.add_argument('--insert-end-mark', action='store_true', help="문서 끝에 '끝' 표시를 자동 삽입")
    parser.add_argument(
        '--doc-type', choices=['plan', 'sihaengmun'], default='plan',
        help='항목체계 최상위 레벨: plan=Ⅰ.(계획서·보고서, 기본) / sihaengmun=1.(대외 시행문, 로마숫자 미사용)',
    )
    parser.add_argument('--list-formats', action='store_true', help='지원 형식 목록 출력')
    parser.add_argument('--preflight', action='store_true', help='HWP COM 실행 가능 여부만 점검하고 종료')
    parser.add_argument('--startup-timeout', type=int, default=45, help='HWP COM 시작 제한 시간(초)')
    parser.add_argument('--diagnose-stages', action='store_true', help='실제 변환 hang 진단용 단계 로그 출력')
    parser.add_argument('--kordoc-home', default=None, help='스캔 PDF OCR용 kordoc-ai 경로 (또는 KORDOC_HOME 환경변수)')
    parser.add_argument('--_preflight-worker', action='store_true', help=argparse.SUPPRESS)
    parser.add_argument('--_preflight-visible', action='store_true', help=argparse.SUPPRESS)
    args = parser.parse_args(argv)

    if args._preflight_worker:
        try:
            print(_run_hwp_preflight_worker(visible=args._preflight_visible))
            return 0
        except Exception as exc:
            print(str(exc), file=sys.stderr)
            return 2

    if args.list_formats:
        print('지원 입력 형식: ' + ', '.join(sorted(SUPPORTED_EXTENSIONS)))
        return 0

    if args.preflight:
        try:
            print(run_hwp_preflight(visible=False, timeout=args.startup_timeout))
            return 0
        except Exception as exc:
            print(f'[FAIL] {exc}', file=sys.stderr)
            return 2

    if not args.files:
        parser.error('변환할 파일 경로가 필요함')
    if args.empty_output_folder and not args.output_dir:
        parser.error('--empty-output-folder는 -o/--output-dir와 함께 사용해야 함')

    global _ALLOW_ROMAN_LEVEL
    _ALLOW_ROMAN_LEVEL = (args.doc_type != 'sihaengmun')  # 정본 §2-1 항목체계 최상위 레벨

    hwp = None
    failures = []
    try:
        prepared_output_dir = prepare_output_dir(args.output_dir, args.empty_output_folder) if args.output_dir else None
        print('HWP 실행 중...')
        try:
            hwp = create_hwp_object(visible=True)
        except Exception as exc:
            print(f'[FAIL] {exc}', file=sys.stderr)
            return 2
        time.sleep(1.5)

        for src_arg in args.files:
            try:
                src_path = as_path(src_arg)
                hwpx_path = build_output_path(src_path, prepared_output_dir)
                print(f'변환 중: {src_path.name} → {hwpx_path.name}')
                diagnose_stage = None
                if args.diagnose_stages:
                    def diagnose_stage(stage):
                        print(f'[diagnose] {stage}', flush=True)

                convert_file(
                    hwp,
                    src_path,
                    hwpx_path,
                    insert_end_mark=args.insert_end_mark,
                    kordoc_home=args.kordoc_home,
                    diagnose_stage=diagnose_stage,
                )
                record_output_file(prepared_output_dir, hwpx_path)
            except Exception as exc:
                failures.append((src_arg, exc))
                print(f'[FAIL] {src_arg}: {exc}', file=sys.stderr)
    finally:
        if hwp is not None:
            hwp.Quit()

    if failures:
        print('\n실패 목록:', file=sys.stderr)
        for src_arg, exc in failures:
            print(f'- {src_arg}: {exc}', file=sys.stderr)
        return 1
    print('\n전체 변환 완료.')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
